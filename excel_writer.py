# excel_writer.py
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from typing import Dict, Any, List

EXCEL_FILE = "lankapropertyweb_land_listings.xlsx"
SUMMARY_SHEET = "summary"
DETAIL_SHEET = "detail"

def get_or_create_wb():
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE)
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        if "Sheet" in wb.sheetnames: wb.remove(wb["Sheet"])
        wb.create_sheet(SUMMARY_SHEET)
        wb.create_sheet(DETAIL_SHEET)
    return wb

def write_row(ws: Worksheet, row: List[Any]):
    ws.append(row)
    for i, _ in enumerate(row, 1):
        ws.column_dimensions[get_column_letter(i)].width = 25

def ensure_headers(ws: Worksheet, headers: List[str]):
    if ws.max_row > 1: return
    if ws.max_row == 1 and all(cell.value is None for cell in ws[1]):
        ws.delete_rows(1)
    ws.append(headers)

def save_summary(ad: Dict[str, Any], headers: List[str]):
    wb = get_or_create_wb()
    ws = wb[SUMMARY_SHEET]
    ensure_headers(ws, headers)
    row = [ad.get(h, "") for h in headers]
    write_row(ws, row)
    wb.save(EXCEL_FILE)

def save_detail(ad: Dict[str, Any], headers: List[str]):
    wb = get_or_create_wb()
    ws = wb[DETAIL_SHEET]
    ensure_headers(ws, headers)
    row = [ad.get(h, "") for h in headers]
    write_row(ws, row)
    wb.save(EXCEL_FILE)